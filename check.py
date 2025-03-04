import os
import time
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright
from rich import print


def save_to_excel(product_title, product_price, product_url, workbook):
    """Saves product details to an Excel file."""
    sheet = workbook.active

    # Find the next available row in the sheet
    next_row = sheet.max_row + 1

    # Write product details into the row
    sheet[f"A{next_row}"] = product_title
    sheet[f"B{next_row}"] = product_price
    sheet[f"C{next_row}"] = product_url

    print(f"Saved: Row {next_row} - {product_title}")


def run(playwright):
    start_url = "https://www.daraz.com.bd/catalog/?q=router"
    browser = playwright.chromium.launch(headless=False)

    # Create a new Excel workbook and add headers
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Products"
    sheet["A1"] = "Product Name"
    sheet["B1"] = "Price"
    sheet["C1"] = "URL"

    # Step 1: Open search results page ONCE
    page = browser.new_page()
    page.goto(start_url)
    print("Opened search results page.")

    # Step 2: Collect all product links
    links = page.locator("a[href*='tp-link']").all()
    product_urls = [link.get_attribute("href") for link in links if link.get_attribute("href")]

    # Step 3: Close the search results page
    page.close()
    print("Closed search results page.")

    # Step 4: Open each product page, extract data, and save to Excel
    for idx, url in enumerate(product_urls):
        full_url = "https:" + url
        print(f"Opening product page: {full_url}")

        page = browser.new_page()

        # Retry mechanism with timeout handling
        try:
            page.goto(full_url, timeout=120000)  # Increase timeout to 2 minutes
            page.wait_for_selector("h1")  # Wait for the product title to appear

            # Extract product details
            title = page.locator("h1").text_content() or "No Title Found"

            # ✅ Wait for price to appear, then extract
            try:
                price_selector = ".pdp-price"  # ✅ Use correct selector (Inspect in DevTools)
                page.wait_for_selector(price_selector, timeout=10000)
                price = page.locator(price_selector).text_content().strip()
            except:
                price = "No Price Found"

            # Save to Excel file (one row per product)
            save_to_excel(title.strip(), price.strip(), full_url, workbook)

        except playwright._impl._errors.TimeoutError:
            print(f"Timeout occurred for {full_url}. Skipping this product.")

        time.sleep(2)  # Keep the page open for a few seconds
        page.close()
        print(f"Closed product page: {full_url}")

    # Save the Excel workbook to a file
    workbook.save("products.xlsx")
    print("Finished scraping all products. Data saved to products.xlsx.")

    browser.close()


if __name__ == "__main__":
    with sync_playwright() as p:
        run(p)
